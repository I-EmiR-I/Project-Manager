from datetime import datetime, timedelta,date
import random
from itertools import groupby
from PyQt5.QtWidgets import QMainWindow, QTabWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget, QApplication
from PyQt5 import uic
from PyQt5.QtGui import QPixmap
from PyQt5.QtCore import Qt
import PyQt5
import openpyxl

class MainWindow(PyQt5.QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        # Load the UI file
        self.main = uic.loadUi("gui.ui", self)
        self.setWindowTitle("PA TECH")
        self.setWindowIcon(PyQt5.QtGui.QIcon('logo.png'))
        self.gui_configurations()
        self.switch_tab(index=2)


    def gui_configurations(self):

        #registro
        self.db = self.findChild(PyQt5.QtWidgets.QTableWidget, "tableWidget")
        self.db.hide()

        #activos
        self.db_active = self.findChild(PyQt5.QtWidgets.QTableWidget,"tableWidget_2")
        self.db_active.hide()

        #terminados
        self.db_term = self.findChild(PyQt5.QtWidgets.QTableWidget,"tableWidget_3")
        self.db_term.hide()

        # set columns size
        for db in [self.db,self.db_active,self.db_term]:
            db.setColumnWidth(0, 100)
            db.setColumnWidth(1, 400)
            db.setColumnWidth(5, 210)

        # buttons command
        self.btn_detalles = self.findChild(PyQt5.QtWidgets.QPushButton,"pushButton_13")
        self.btn_detalles.clicked.connect(self.AddGastoUI)

        self.btn_amp_garantia = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_14")
        self.btn_amp_garantia.clicked.connect(self.end_project)

        self.btn_remover_garantia = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_15")
        self.btn_remover_garantia.clicked.connect(lambda:self.end_project(m=True))
        self.btn_ver = self.findChild(PyQt5.QtWidgets.QPushButton,"pushButton_10")

        self.btn_gasto = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_11")
        self.btn_gasto.clicked.connect(self.AddGastoUI)

        self.btn_terminar = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_12")
        self.btn_terminar.clicked.connect(self.end_project)

        self.btn_agregar = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_7")
        self.btn_agregar.clicked.connect(self.AddProjectUI)

        self.btn_editar = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_8")
        self.btn_editar.clicked.connect(self.EditProjectUI)

        self.btn_iniciar = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_9")
        self.btn_iniciar.clicked.connect(self.active_project)

        # navigation buttons
        self.btn_registro = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton")
        self.btn_registro.clicked.connect(lambda:self.switch_tab(2))

        self.btn_activo = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_2")
        self.btn_activo.clicked.connect(lambda:self.switch_tab(1))

        self.btn_terminado = self.findChild(PyQt5.QtWidgets.QPushButton, "pushButton_3")
        self.btn_terminado.clicked.connect(lambda:self.switch_tab(0))

    def currentRow(self):
        items = []
        for x in range(0, int(self.dbs.columnCount())):
            items.append(self.dbs.item(self.dbs.currentRow(), x).text())
        return items

    def AddGastoUI(self):

        #window load and settings
        try:#get item
            self.items = self.currentRow()
        except:
            return

        self.gasto_ui = uic.loadUi("AgregarGasto.ui",self)
        self.setWindowTitle("Agregar gasto")
        self.gasto_ui.show()

        #text default
        self.lineEdit.setText(self.items[1])
        self.lineEdit_2.setText(self.items[2])
        self.lineEdit_2.setReadOnly(1)
        self.lineEdit_3.setText(self.items[0])
        self.lineEdit_3.setReadOnly(1)
        self.lineEdit_4.setText(str(date.today()))
        self.lineEdit_4.setReadOnly(1)
        #material
        self.lineEdit_5.setText("")
        #costo del material
        self.lineEdit_6.setText("0")
        self.lineEdit_7.setReadOnly(1)

        #botones

        #agregar comprobante
        self.pushButton.clicked.connect(lambda : self.select_image(self.items,int(self.lineEdit_6.text()),self.lineEdit_5.text(),int(self.lineEdit_8.text())))
        #cancelar
        self.pushButton_2.clicked.connect(lambda :self.cancelGasto())
        self.pushButton_3.clicked.connect(lambda: self.add_back(self.gasto_ui))
        #connect db to every event
        self.tableWidget.setSelectionBehavior(PyQt5.QtWidgets.QAbstractItemView.SelectRows)  # enable row selection
        self.tableWidget.itemSelectionChanged.connect(self.on_selection_changed)



        #populate gastos
        self.populateGastos(self.items)

    def on_selection_changed(self):
           try:
                # get the selected rows
                items = self.tableWidget.selectedItems()
                item = items[-1].text()
                self.material = items[0].text()
                pixmap = QPixmap(item)
                self.label_7 = self.findChild(PyQt5.QtWidgets.QLabel, "label_7")
                label_width = self.label_7.width()
                label_height = self.label_7.height()
                scaled_pixmap = pixmap.scaled(label_width, label_height, PyQt5.QtCore.Qt.KeepAspectRatio)
                self.label_7.setPixmap(scaled_pixmap)
           except:
                pass
    def AddProjectUI(self):
        add_project_ui = uic.loadUi("gui2.ui",self)
        # Show the UI
        add_project_ui.show()
        self.setWindowTitle("Agregar proyectos")

        self.pushButton.clicked.connect(lambda :self.add_data(add_project_ui))
        self.pushButton_2.clicked.connect(lambda: self.add_back(add_project_ui))

        self.lineEdit_4.setText(str(date.today()))
        self.lineEdit_4.setReadOnly(1)
        self.lineEdit_3.setText(str(self.id+1))
        self.lineEdit_3.setReadOnly(1)


    def EditProjectUI(self):
        try:#get item
            items = []
            for x in range(0,int(self.dbs.columnCount())):
                items.append(self.dbs.item(self.dbs.currentRow(),x).text())
        except:
            return
        #build gui
        edit_project_ui = uic.loadUi("gui2.ui", self)
        # Show the UI
        edit_project_ui.show()

        self.setWindowTitle("Editar proyectos")

        #set button functions
        self.pushButton.clicked.connect(lambda: self.add_data(edit_project_ui))
        self.pushButton_2.clicked.connect(lambda: self.add_back(edit_project_ui))
        #set items
        self.lineEdit_5.setText(items[3])#cliente
        self.lineEdit_4.setText(items[4])#fecha
        self.lineEdit_4.setReadOnly(1)
        self.lineEdit_3.setText(str(items[0]))#codigo
        self.lineEdit_3.setReadOnly(1)
        self.lineEdit_2.setText(items[2])#direccion
        self.lineEdit.setText(items[1])#nombre
        self.plainTextEdit.setPlainText(self.searchCode(int(items[0])))

    def cancelGasto(self):
        # Load the Excel workbook
        workbook = openpyxl.load_workbook("Proyectos.xlsx")
        worksheet = workbook.active
        lista = []

        selected_item = self.tableWidget.item(self.tableWidget.currentRow(), 0)
        if selected_item is not None:
            value = selected_item.data(Qt.UserRole)
            print("Selected value:", value)
        else:
            return 0

        # Define the ID you want to extract data for
        id_to_extract = value
        item = self.tableWidget.item(self.tableWidget.currentRow(), 1)
        print(f"costo {int(item.text())}")
        if value != None:
            # Iterate through each row in the worksheet, starting from row 2 (assuming the first row contains column headers)
            n = 2
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # Check if the ID in the current row matches the ID to extract
                if row[11] == id_to_extract:
                    costo = row[7]
                    # If it matches, extract the values in the current row
                    worksheet["H"+str(n)] = 0
                n+=1
            workbook.save("Proyectos.xlsx")
            workbook.close()
            self.tableWidget.removeRow(self.tableWidget.currentRow())
            self.costo = self.costo - costo
            print(costo)
            #print((int(self.costo)) - (int(item.text())) )
            self.lineEdit_7.setText(str(self.costo))
    def populateGastos(self,items):
        table = self.tableWidget

        # Abrir el archivo de Excel
        workbook = openpyxl.load_workbook('Proyectos.xlsx')

        # Seleccionar la hoja de Excel
        sheet = workbook.active

        # Leer las filas y crear una lista de diccionarios
        rows = []
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            rows.append({
                'Codigo': values[0],
                'Nombre': values[1],
                'Direccion': values[2],
                'Fecha': values[3],
                'Cliente':values[4],
                'Retorno':values[5],
                'Precio': values[7],
                'Material':values[9],
                'Comprobante':values[8],
                'Identificador':values[11]
            })
        # Agrupar las filas por el valor de la columna "Codigo"
        rows_sorted = sorted(rows, key=lambda x: x['Codigo'])
        groups = groupby(rows_sorted, key=lambda x: x['Codigo'])

        # Crear una lista de listas con los diccionarios de las filas agrupadas
        proyectos = []
        for key, group in groups:
            proyectos.append(list(group))

        # Agregar las filas a la tabla
        self.datos = []
        costo = 0
        for grupo in proyectos:
            codigo = grupo[0]['Codigo']
            for proyecto in grupo:
                if str(codigo) == str(items[0]):
                    if proyecto["Precio"] != 0:
                        row = table.rowCount()
                        table.insertRow(row)
                        table.setItem(row, 0, PyQt5.QtWidgets.QTableWidgetItem(proyecto['Material']))
                        table.setItem(row, 1, PyQt5.QtWidgets.QTableWidgetItem(str(proyecto['Precio'])))
                        table.setItem(row, 2, PyQt5.QtWidgets.QTableWidgetItem(proyecto['Comprobante']))
                        self.item = table.item(row,0)
                        self.item.setData(Qt.UserRole,proyecto["Identificador"])
                        self.datos.append((proyecto["Material"],proyecto["Codigo"],proyecto["Identificador"]))
                        costo += int(proyecto["Precio"])
        print(f"Datos: {self.datos}")
        self.costo = costo
        print(f"COSTO TOTAL DEL PROYECTO {self.costo}")

        #set costo
        self.lineEdit_7.setText(str(self.costo))
        self.lineEdit_8.setText(str(proyecto["Retorno"]))
    def select_image(self,items,precio,material,retorno):
        # Open the file dialog
        filename = None
        if precio != 0:
            filename, _ = PyQt5.QtWidgets.QFileDialog.getOpenFileName(self, 'Select Image', '',
                                                                      'Images (*.png *.jpg *.jpeg)')

        if 1:
            # Load the workbook
            workbook = openpyxl.load_workbook("Proyectos.xlsx")
            worksheet = workbook.active

            # Write the URL to the cell
            comprobante = filename
            id = items[0]
            cliente = self.searchCode(id,x=5)
            nombre = items[1]
            direccion = items[2]
            fecha = items[3]

            #aqui deben de ir el precio y cobranza del material y no del proyecto
            precio = precio
            #nombre, direccion, fecha, cliente, retorno = 0, desc = None, precio = 0, comprobante = None, material = None, estado = None, codigo = None
            self.add_data_to_excel(nombre, direccion, fecha, cliente,codigo=int(id),comprobante=comprobante,precio=precio,retorno=retorno,estado="Activo",material=material)

            #add new gasto to the table
            table = self.tableWidget
            if precio != 0:
                row = table.rowCount()
                table.insertRow(row)
                table.setItem(row, 0, PyQt5.QtWidgets.QTableWidgetItem(material))
                table.setItem(row, 1, PyQt5.QtWidgets.QTableWidgetItem(str(precio)))
                table.setItem(row, 2, PyQt5.QtWidgets.QTableWidgetItem(comprobante))
                self.item = table.item(row, 0)
                self.item.setData(Qt.UserRole, self.identificador)
                #here
                self.datos.append((material,id,self.identificador))
                self.costo += precio

            #update costo
            self.lineEdit_7.setText(str(self.costo))

            # material
            self.lineEdit_5.setText("")
            # costo del material
            self.lineEdit_6.setText("0")


    def add_data(self,gui):
        self.get_input_values_from_add()
        gui.destroy()
        self.__init__()
        self.main.show()

    def add_back(self,gui):
        gui.destroy()
        self.__init__()
        self.main.show()

    def searchCode(self,id,x=6):

        # Load the Excel workbook
        workbook = openpyxl.load_workbook("Proyectos.xlsx")
        worksheet = workbook.active
        lista = []

        # Define the ID you want to extract data for
        id_to_extract = int(id)

        # Iterate through each row in the worksheet, starting from row 2 (assuming the first row contains column headers)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Check if the ID in the current row matches the ID to extract
            if row[0] == id_to_extract:
                # If it matches, extract the values in the current row
                col1_value = row[x]
                print(col1_value)
                if x == 9:
                    lista.append(col1_value)
        if x ==9:
            return lista
        return col1_value

    #this function add and edits projects
    def get_input_values_from_add(self):
        nombre = self.lineEdit.text()
        direccion = self.lineEdit_2.text()
        codigo = int(self.lineEdit_3.text())
        fecha = self.lineEdit_4.text()
        estado = "Registro"
        descripcion = self.plainTextEdit.toPlainText()
        cliente = self.lineEdit_5.text()
        self.add_data_to_excel(nombre,direccion,fecha,cliente,retorno=0,desc=descripcion,estado=estado,codigo=codigo)

    def end_project(self,m= False):
        print("asd")
        items = []
        for x in range(0, int(self.dbs.columnCount())):
            items.append(self.dbs.item(self.dbs.currentRow(), x).text())
        codigo = items[0]
        print(codigo)

        workbook = openpyxl.load_workbook("Proyectos.xlsx")
        worksheet = workbook.active

        # iterate over rows to find ID
        for row in worksheet.iter_rows(min_row=2, max_col=1):
            if row[0].value == int(codigo):
                # update value of specific cell
                worksheet.cell(row=row[0].row, column=11).value = 'Terminado'
                if m != False:
                    worksheet.cell(row=row[0].row, column=4).value = "Sin Garantia"
                    print("falso", m)
                else:
                    worksheet.cell(row=row[0].row, column=4).value = (str(date.today() + timedelta(days=365)))
                    print((str(date.today() + timedelta(days=365))))
        # save changes to workbook
        workbook.save('Proyectos.xlsx')
        self.switch_tab(0)
    def active_project(self):
        items = []
        for x in range(0, int(self.dbs.columnCount())):
            items.append(self.dbs.item(self.dbs.currentRow(), x).text())
        codigo = items[0]
        print(codigo)

        workbook = openpyxl.load_workbook("Proyectos.xlsx")
        worksheet = workbook.active

        # iterate over rows to find ID
        for row in worksheet.iter_rows(min_row=2, max_col=1):
            if row[0].value == int(codigo):
                # update value of specific cell
                worksheet.cell(row=row[0].row, column=11).value = 'Activo'
                worksheet.cell(row=row[0].row, column=4).value = str(date.today())
        # save changes to workbook
        workbook.save('Proyectos.xlsx')
        self.switch_tab(1)
    def load_data_into_table(self, data):
        # Set the number of rows and columns
        db = self.dbs
        db.setRowCount(0)


        # Set the horizontal header labels
        header = {self.db:["Codigo", "Nombre", "Direccion","Cliente","Fecha de registro","Cobranza"],
                  self.db_active:["Codigo","Nombre","Direccion","Fecha de inicio","Costo","Cobranza"],
                  self.db_term:["Codigo","Nombre","Direccion","Fecha de garantia","Costo","Cobranza"]}

        condition = {self.db:"Registro",self.db_active:"Activo",self.db_term:"Terminado"}

        db.setHorizontalHeaderLabels(header[db])
        db.setColumnCount(len(header[db]))

        #set size
        try:

            try:
                db.setColumnWidth(header[db].index("Direccion"),330)
                db.setColumnWidth(header[db].index("Nombre"),330)
                db.setColumnWidth(header[db].index("Fecha de registro"), 230)
                db.setColumnWidth(header[db].index("Cliente"), 330)
                db.setColumnWidth(header[db].index("Cobranza"), 175)
                db.setColumnWidth(header[db].index("Codigo"), 90)
                fecha = "Fecha de registro"
            except:
                db.setColumnWidth(header[db].index("Direccion"),345)
                db.setColumnWidth(header[db].index("Nombre"),345)
                db.setColumnWidth(header[db].index("Fecha de inicio"), 233)
                db.setColumnWidth(header[db].index("Cobranza"), 233)
                db.setColumnWidth(header[db].index("Costo"), 233)
                fecha = "Fecha de inicio"
        except:
            db.setColumnWidth(header[db].index("Direccion"), 345)
            db.setColumnWidth(header[db].index("Nombre"), 345)
            db.setColumnWidth(header[db].index("Fecha de garantia"), 233)
            db.setColumnWidth(header[db].index("Cobranza"), 233)
            db.setColumnWidth(header[db].index("Costo"), 233)
            fecha = "Fecha de garantia"

        print(data)
        # Add the data to the table
        rows = 0
        for row in range(len(data["codigo"])):
            if condition[db] == str(data["estado"][row]):
                rows+=1
                db.setRowCount(rows)
                db.setItem(rows-1, 0, PyQt5.QtWidgets.QTableWidgetItem(str(data["codigo"][row])))
                db.setItem(rows-1, 1, PyQt5.QtWidgets.QTableWidgetItem(str(data["nombre"][row])))
                db.setItem(rows-1, 2, PyQt5.QtWidgets.QTableWidgetItem(str(data["direccion"][row])))
                db.setItem(rows-1, header[db].index("Cobranza"), PyQt5.QtWidgets.QTableWidgetItem(str(data["retorno"][row])))
                db.setItem(rows - 1, header[db].index(fecha),PyQt5.QtWidgets.QTableWidgetItem(str(data["fecha"][row])))
                if fecha == "Fecha de registro":
                    db.setItem(rows - 1, header[db].index("Cliente"),
                               PyQt5.QtWidgets.QTableWidgetItem(str(data["cliente"][row])))
                else:
                    db.setItem(rows - 1, header[db].index("Costo"),
                               PyQt5.QtWidgets.QTableWidgetItem(str(data["precio"][row])))
    def load_data_from_excel(self):
        # Load the data from the Excel file
        workbook = openpyxl.load_workbook("Proyectos.xlsx")
        worksheet = workbook.active

        codigo = []
        nombre = []
        direccion = []
        fecha = []
        cliente = []
        retorno = []
        desc = []
        precio = []
        comprobante = []
        material = []
        estado = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] in codigo and row[0] != None:
                index = codigo.index(row[0])
                nombre[index] = row[1]
                direccion[index] = row[2]
                fecha[index] = row[3]
                cliente[index] = row[4]
                retorno[index] = row[5]
                desc[index] = row[6]
                precio[index] += row[7]
                comprobante[index] = row[8]
                material[index] = row[9]
                estado[index] = row[10]
            else:
                codigo.append(row[0])
                nombre.append(row[1])
                direccion.append(row[2])
                fecha.append(row[3])
                cliente.append(row[4])
                retorno.append(row[5])
                desc.append(row[6])
                precio.append(row[7])
                comprobante.append(row[8])
                material.append(row[9])
                estado.append(row[10])

        # Return the data as a dictionary
        data = {
            "codigo": codigo,
            "nombre": nombre,
            "direccion": direccion,
            "fecha": fecha,
            "cliente": cliente,
            "retorno": retorno,
            "desc": desc,
            "precio": precio,
            "comprobante": comprobante,
            "material": material,
            "estado": estado
        }
        print(data)
        self.id = max(data["codigo"])
        return data

    def add_data_to_excel(self,nombre, direccion, fecha, cliente, retorno=0, desc=None, precio=0, comprobante=None, material=None, estado=None,codigo=None):
        # Load the workbook
        workbook = openpyxl.load_workbook("Proyectos.xlsx")

        # Select the active worksheet
        worksheet = workbook.active

        # Get the next available row
        next_row = worksheet.max_row + 1
        self.identificador = next_row + 1000
        # Get the next available codigo value
        codigo_column = worksheet["A"]
        codigo_values = [cell.value for cell in codigo_column[1:]]
        max_codigo = max(codigo_values) if codigo_values else 0
        next_codigo = max_codigo + 1

        # Add the data to the next row
        if codigo == None:
            worksheet.cell(row=next_row, column=1, value=next_codigo)
        else:
            worksheet.cell(row=next_row, column=1, value=codigo)
        worksheet.cell(row=next_row, column=2, value=nombre)
        worksheet.cell(row=next_row, column=3, value=direccion)
        worksheet.cell(row=next_row, column=4, value=fecha)
        worksheet.cell(row=next_row, column=5, value=cliente)
        worksheet.cell(row=next_row, column=6, value=retorno)
        worksheet.cell(row=next_row, column=7, value=desc)
        worksheet.cell(row=next_row, column=8, value=precio)
        worksheet.cell(row=next_row, column=9, value=comprobante)
        worksheet.cell(row=next_row, column=10, value=material)
        worksheet.cell(row=next_row, column=11, value=estado)
        worksheet.cell(row=next_row, column=12, value=self.identificador)

        # Save the changes
        workbook.save("Proyectos.xlsx")

    def switch_tab(self, index):
        #hide tables
        dbs = [self.db_term,self.db_active,self.db]
        for x in dbs:
            x.hide()

        #hide buttons
        btn = [self.btn_detalles, self.btn_amp_garantia, self.btn_remover_garantia, self.btn_ver, self.btn_gasto,
               self.btn_terminar, self.btn_agregar, self.btn_editar, self.btn_iniciar]
        for x in btn:
            x.hide()

        #show table according to the given index
        self.dbs = dbs[index]
        dbs[index].show()
        btn[index*3].show()
        btn[index*3+1].show()
        btn[index*3+2].show()

        data = self.load_data_from_excel()
        self.load_data_into_table(data)

if __name__ == "__main__":
    import sys
    app = PyQt5.QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())

